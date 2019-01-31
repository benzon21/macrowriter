import pyodbc
from openpyxl import Workbook, load_workbook

def accesstoexcel(savename,sql):
	print("Creating Workbook")
	wb = Workbook()
	wsh = wb.get_sheet_by_name("Sheet")
	wsh.title = "Material Data"
	
	print("Connect to Database")
	conn_str = (
		r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
		r'DBQ=MATERIAL_DATABASE.accdb;'
		)
		
	cnxn = pyodbc.connect(conn_str)

	crsr = cnxn.cursor()

	crsr.execute(sql) 

	rows = crsr.fetchall()

	total_rows = len(rows)

	current_row = 1
	current_column = 1

	titles = crsr.description

	for title in titles:
		wsh.cell(row=current_row,column=current_column).value = title[0]
		current_column += 1

	next_column = 1
	next_row = 2

	row_array = 0


	for row in rows:
		num_row = len(row)
		for num in range(0,num_row):
			wsh.cell(row=next_row,column=next_column).value = rows[row_array][num]
			next_column += 1
		next_row += 1
		row_array += 1
		next_column = 1
		
	crsr.close()
	cnxn.close()
	wb.save(savename)
	wb.close()
	print("Saved.")

def multiplemix(mixes):
	for mix in mixes:
		savename = "Materials\\%s.xlsx" % (mix)
		sql = "SELECT * FROM MATERIAL_TABLE WHERE MATERIAL = '%s'"% (mix)
		accesstoexcel(savename,sql)
	
def categories(sql):
	print("Connect to Database")
	conn_str = (
		r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
		r'DBQ=MATERIAL_DATABASE.accdb;'
		)
		
	cnxn = pyodbc.connect(conn_str)
	crsr = cnxn.cursor()
	crsr.execute(sql)
	
	titles = crsr.description
	arr = []
	for title in titles:
		arr.append(str(title[0]))
	
	crsr.close()
	cnxn.close()
	return arr

def limits(excel):
	limit_array = []
	limit_worbook = load_workbook(filename=excel , read_only=True)
	limit_ws = limit_worbook['Batch Data']
	for num in range(2,28,2):
		group = []
		group.append(limit_ws.cell(row=2,column=num).value)
		group.append(limit_ws.cell(row=2,column=num+1).value)
		limit_array.append(group)
	return limit_array

def macro(file_name,save_name,sql,mix_type):
	new_arr = categories(sql)[1:14]
	new_text = open(file_name,'w')
	new_text.write("GMACRO \n")
	my_arr = limits(save_name)
	start_num = 0
	for item in new_arr:
		new_value = my_arr[start_num]
		new_text.write("\nLayout. \n")
		new_text.write("IChart '%s';\n"%(item))
		new_text.write("Title '%s %s';\n"%(mix_type, item))
		new_text.write("Reference 2 %s %s;\n"%(new_value[0], new_value[1]))
		new_text.write("LABEL 'LSL= %s' 'USL= %s'.\n"%(new_value[0], new_value[1]))
		new_text.write("Endlayout. \n")
		if(start_num < len(my_arr)):
			start_num += 1
	new_text.write("\nENDMACRO")
	new_text.close()	


